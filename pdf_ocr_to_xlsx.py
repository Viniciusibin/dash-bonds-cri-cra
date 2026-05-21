#!/usr/bin/env python3
"""
OCR de PDF para XLSX.
Extrai texto de cada página via Tesseract e salva em planilha estruturada.

Dependências:
    pip install pdf2image pytesseract openpyxl Pillow

Também requer:
    - Tesseract OCR: https://github.com/UB-Mannheim/tesseract/wiki
      (marcar "Portuguese" no instalador ou baixar por.traineddata)
    - Poppler para pdf2image no Windows:
      https://github.com/oschwartz10612/poppler-windows/releases
      (extrair e adicionar a pasta bin ao PATH)
"""

import sys
import argparse
from pathlib import Path


TESSERACT_DEFAULT = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
POPPLER_DEFAULT = r"C:\poppler\poppler-26.02.0\Library\bin"


def check_deps():
    missing = []
    for mod, pkg in [("pdf2image", "pdf2image"), ("pytesseract", "pytesseract"),
                     ("PIL", "Pillow"), ("openpyxl", "openpyxl")]:
        try:
            __import__(mod)
        except ImportError:
            missing.append(pkg)
    if missing:
        print("Instale as dependências:\n  pip install", " ".join(missing))
        sys.exit(1)


def ocr_page(image, tesseract_cmd, lang):
    """Retorna dados OCR com posição de cada palavra."""
    import pytesseract
    pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
    return pytesseract.image_to_data(
        image,
        lang=lang,
        config="--psm 6 --oem 3",
        output_type=pytesseract.Output.DICT,
    )


def reconstruct_lines(ocr_data, row_tol=12):
    """Agrupa palavras por linha usando coordenada Y."""
    words = [
        {"text": ocr_data["text"][i].strip(),
         "left": ocr_data["left"][i],
         "top":  ocr_data["top"][i]}
        for i in range(len(ocr_data["text"]))
        if ocr_data["text"][i].strip() and int(ocr_data["conf"][i]) > 20
    ]
    if not words:
        return []

    words.sort(key=lambda w: w["top"])
    buckets, current = [], [words[0]]
    for w in words[1:]:
        if abs(w["top"] - current[-1]["top"]) <= row_tol:
            current.append(w)
        else:
            buckets.append(sorted(current, key=lambda x: x["left"]))
            current = [w]
    buckets.append(sorted(current, key=lambda x: x["left"]))

    return [" ".join(w["text"] for w in bucket) for bucket in buckets]


def save_xlsx(pages, output_path):
    """Salva uma aba por página + aba consolidada."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="2F5496")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    PAGE_FILL   = PatternFill("solid", fgColor="D9E1F2")
    PAGE_FONT   = Font(bold=True, size=10)

    wb = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = "Texto Completo"

    # Cabeçalho da aba consolidada
    ws_all.cell(1, 1, "Pág").fill  = HEADER_FILL
    ws_all.cell(1, 1).font         = HEADER_FONT
    ws_all.cell(1, 2, "Linha").fill = HEADER_FILL
    ws_all.cell(1, 2).font          = HEADER_FONT
    ws_all.cell(1, 3, "Texto").fill = HEADER_FILL
    ws_all.cell(1, 3).font          = HEADER_FONT
    ws_all.column_dimensions["C"].width = 130

    all_row = 2
    for page_num, lines in pages:
        for line_num, line in enumerate(lines, 1):
            ws_all.cell(all_row, 1, page_num)
            ws_all.cell(all_row, 2, line_num)
            ws_all.cell(all_row, 3, line)
            all_row += 1

        # Aba individual por página
        ws = wb.create_sheet(f"Pag {page_num}")
        ws.cell(1, 1, f"Página {page_num}").fill = PAGE_FILL
        ws.cell(1, 1).font = PAGE_FONT
        ws.column_dimensions["A"].width = 130
        for r, line in enumerate(lines, 2):
            ws.cell(r, 1, line)

    wb.save(output_path)
    print(f"XLSX salvo em: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="OCR de PDF → XLSX")
    parser.add_argument("pdf", nargs="?",
                        default="05 - Plano de Recuperação Judicial_parte_001.pdf")
    parser.add_argument("--lang", default="por+eng",
                        help="Idiomas Tesseract (padrão: por+eng)")
    parser.add_argument("--dpi", type=int, default=300,
                        help="DPI de renderização (padrão: 300)")
    parser.add_argument("--tesseract", default=TESSERACT_DEFAULT,
                        help="Caminho do executável tesseract.exe")
    parser.add_argument("--poppler", default=None,
                        help="Pasta bin do Poppler (se não estiver no PATH)")
    parser.add_argument("--raw-txt", action="store_true",
                        help="Também salvar texto bruto em .txt")
    args = parser.parse_args()

    check_deps()

    from pdf2image import convert_from_path

    pdf_path = Path(args.pdf)
    if not pdf_path.exists():
        print(f"Arquivo não encontrado: {pdf_path}")
        sys.exit(1)

    tesseract = args.tesseract
    if not Path(tesseract).exists():
        # Tenta encontrar no PATH
        import shutil
        found = shutil.which("tesseract")
        if found:
            tesseract = found
        else:
            print(f"Tesseract não encontrado em: {tesseract}")
            print("Instale em https://github.com/UB-Mannheim/tesseract/wiki")
            sys.exit(1)

    poppler_path = args.poppler
    if poppler_path is None and Path(POPPLER_DEFAULT).exists():
        poppler_path = POPPLER_DEFAULT

    print(f"PDF     : {pdf_path}")
    print(f"DPI     : {args.dpi}")
    print(f"Idioma  : {args.lang}")
    print(f"Tesseract: {tesseract}")
    print()

    print("Convertendo páginas do PDF para imagem...")
    convert_kwargs = {"dpi": args.dpi}
    if poppler_path:
        convert_kwargs["poppler_path"] = poppler_path

    images = convert_from_path(str(pdf_path), **convert_kwargs)
    print(f"Páginas encontradas: {len(images)}\n")

    pages = []
    raw_text = []

    for i, img in enumerate(images, 1):
        print(f"OCR página {i}/{len(images)}...", end=" ", flush=True)
        data = ocr_page(img, tesseract, args.lang)
        lines = reconstruct_lines(data)
        pages.append((i, lines))
        print(f"{len(lines)} linhas")

        if args.raw_txt:
            raw_text.append(f"\n{'='*60}\nPÁGINA {i}\n{'='*60}")
            raw_text.extend(lines)

    if args.raw_txt:
        txt_path = pdf_path.with_suffix(".txt")
        txt_path.write_text("\n".join(raw_text), encoding="utf-8")
        print(f"\nTexto bruto: {txt_path}")

    output_path = pdf_path.with_suffix(".xlsx")
    save_xlsx(pages, output_path)
    print("\nConcluído!")


if __name__ == "__main__":
    main()
