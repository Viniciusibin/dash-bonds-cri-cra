import os
import csv
import re
from datetime import datetime, date

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

NULL_VALUES = {"--", "N/D", "n/d", "nd", "", None}
DATA_START_ROW = 9   # Row index where debenture data begins (0-based)
HEADER_ROW = 7       # Row index of column headers


def _is_null(val):
    return val is None or str(val).strip() in NULL_VALUES


def _parse_num(val):
    if _is_null(val):
        return None
    s = str(val).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _parse_date_str(val):
    """Return YYYY-MM-DD string from various date representations."""
    if _is_null(val):
        return None
    s = str(val).strip()
    # Already ISO
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    # DD/MM/YYYY
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    return s


def _xlrd_date(wb, val):
    """Convert xlrd serial date to YYYY-MM-DD string."""
    try:
        t = xlrd.xldate_as_datetime(float(val), wb.datemode)
        return t.strftime("%Y-%m-%d")
    except Exception:
        return str(val)


def _extract_date_from_filename(filename):
    """Extract YYYY-MM-DD from filenames like debenture-15-05.xls."""
    name = os.path.basename(filename)
    m = re.search(r"(\d{1,2})[-_](\d{2})(?:[-_](\d{2,4}))?", name)
    if m:
        day = m.group(1).zfill(2)
        month = m.group(2).zfill(2)
        year_raw = m.group(3)
        if year_raw:
            year = year_raw if len(year_raw) == 4 else "20" + year_raw
        else:
            year = str(datetime.now().year)
        return f"{year}-{month}-{day}"
    return None


_INVALID_CODE_PATTERN = re.compile(r"[\s\(\)\*\#\.]")

def _is_valid_code(codigo):
    """Debenture/CRI/CRA codes are compact alphanumeric strings, no spaces/symbols."""
    if not codigo or len(codigo) < 4 or len(codigo) > 20:
        return False
    if _INVALID_CODE_PATTERN.search(codigo):
        return False
    return True


def _row_to_debenture(row_vals):
    """Map a list of cell values (15 cols) to a debenture dict."""
    def v(idx):
        return row_vals[idx] if idx < len(row_vals) else None

    codigo = str(v(0)).strip() if not _is_null(v(0)) else None
    if not _is_valid_code(codigo):
        return None

    return {
        "codigo":         codigo,
        "nome":           str(v(1)).strip() if not _is_null(v(1)) else None,
        "vencimento":     str(v(2)).strip() if not _is_null(v(2)) else None,
        "indice":         str(v(3)).strip() if not _is_null(v(3)) else None,
        "taxaCompra":     _parse_num(v(4)),
        "taxaVenda":      _parse_num(v(5)),
        "taxaIndicativa": _parse_num(v(6)),
        "desvioPadrao":   _parse_num(v(7)),
        "intervaloMin":   _parse_num(v(8)),
        "intervaloMax":   _parse_num(v(9)),
        "pu":             _parse_num(v(10)),
        "puPar":          _parse_num(v(11)),
        "duration":       _parse_num(v(12)),
    }


def parse_debentures_xls(filepath):
    """Parse old .xls debentures file. Returns (date_str, [rows])."""
    if not HAS_XLRD:
        raise RuntimeError("xlrd not installed")
    wb = xlrd.open_workbook(filepath, encoding_override="latin-1")
    all_rows = []
    ref_date = None

    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        if ws.nrows <= DATA_START_ROW:
            continue

        # Extract reference date from row 3, col 1
        if ref_date is None:
            date_cell = ws.cell_value(3, 1)
            if date_cell and str(date_cell).strip():
                try:
                    ref_date = _xlrd_date(wb, date_cell)
                except Exception:
                    ref_date = _parse_date_str(date_cell)

        for r in range(DATA_START_ROW, ws.nrows):
            row_vals = [ws.cell_value(r, c) for c in range(ws.ncols)]
            row = _row_to_debenture(row_vals)
            if row:
                row["sheet"] = sheet_name
                all_rows.append(row)

    if ref_date is None:
        ref_date = _extract_date_from_filename(filepath)
    return ref_date, all_rows


def parse_debentures_xlsx(filepath):
    """Parse .xlsx debentures file. Returns (date_str, [rows])."""
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl not installed")
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    all_rows = []
    ref_date = None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_list = list(ws.iter_rows(values_only=True))
        if len(rows_list) <= DATA_START_ROW:
            continue

        # Extract reference date from row 3, col 1
        if ref_date is None:
            date_val = rows_list[3][1] if len(rows_list[3]) > 1 else None
            if date_val is not None:
                if isinstance(date_val, (datetime, date)):
                    ref_date = date_val.strftime("%Y-%m-%d")
                else:
                    ref_date = _parse_date_str(date_val)

        for row_vals in rows_list[DATA_START_ROW:]:
            row = _row_to_debenture(list(row_vals))
            if row:
                row["sheet"] = sheet_name
                all_rows.append(row)

    wb.close()
    if ref_date is None:
        ref_date = _extract_date_from_filename(filepath)
    return ref_date, all_rows


def parse_debentures(filepath):
    """Auto-detect format and parse debentures file."""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".xls":
        return parse_debentures_xls(filepath)
    elif ext in (".xlsx", ".xlsm"):
        return parse_debentures_xlsx(filepath)
    raise ValueError(f"Unsupported debentures format: {ext}")


def parse_cricra_csv(filepath):
    """Parse CRI/CRA CSV. Returns (date_str, [rows])."""
    # Try encodings
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            with open(filepath, "r", encoding=enc, errors="strict") as f:
                sample = f.read(512)
            encoding = enc
            break
        except UnicodeDecodeError:
            continue
    else:
        encoding = "latin-1"

    rows = []
    ref_date = None

    with open(filepath, "r", encoding=encoding, errors="replace") as f:
        content = f.read()

    lines = content.splitlines()
    if not lines:
        return None, []

    # Detect delimiter
    header_line = lines[0]
    delimiter = ";" if header_line.count(";") > header_line.count(",") else ","

    reader = csv.reader(lines, delimiter=delimiter)
    headers = None

    def find_col(headers, candidates):
        for name in candidates:
            for i, h in enumerate(headers):
                if name.lower() in h.lower():
                    return i
        return -1

    for i, row in enumerate(reader):
        row = [c.strip() for c in row]
        if i == 0:
            headers = row
            continue
        if not headers or len(row) < 6:
            continue

        # Build field map once
        if i == 1:
            col = lambda *names: find_col(headers, list(names))
            idx = {
                "dataRef":        col("Data de Refer", "Data Ref"),
                "riscoCredito":   col("Risco de Cr", "Risco Cr"),
                "emissor":        col("Emissor"),
                "serie":          col("érie", "Serie", "rie"),
                "emissao":        col("miss", "Emiss"),
                "codigo":         col("digo", "Codigo"),
                "vencimento":     col("Vencimento"),
                "indice":         col("ndice", "Indice", "Corre"),
                "taxaCompra":     col("Taxa Compra"),
                "taxaVenda":      col("Taxa Venda"),
                "taxaIndicativa": col("Taxa Indicativa"),
                "desvioPadrao":   col("Desvio"),
                "pu":             col("PU"),
                "puPar":          col("PU Par", "% PU", "% VNE"),
                "duration":       col("Duration"),
            }
            # Fix ambiguous: if "Taxa Compra" and "Taxa Venda" map to same col
            # because "Taxa Indicativa" was matched first, adjust
            # PU col must not overlap with puPar
            if idx["pu"] == idx["puPar"] and idx["pu"] != -1:
                # Find the one that literally says "% PU" or "% VNE"
                for j, h in enumerate(headers):
                    if ("% PU" in h or "% VNE" in h) and j != idx["pu"]:
                        idx["puPar"] = j
                        break

        def g(key):
            i_ = idx.get(key, -1)
            return row[i_] if i_ != -1 and i_ < len(row) else ""

        raw_date = g("dataRef")
        if raw_date and not ref_date:
            ref_date = _parse_date_str(raw_date)

        codigo = g("codigo")
        if not _is_valid_code(codigo):
            continue

        def br_num(val):
            if _is_null(val):
                return None
            s = str(val).strip().replace(".", "").replace(",", ".")
            try:
                return float(s)
            except ValueError:
                return None

        rows.append({
            "dataRef":        raw_date,
            "riscoCredito":   g("riscoCredito"),
            "emissor":        g("emissor"),
            "serie":          g("serie"),
            "codigo":         codigo,
            "vencimento":     g("vencimento"),
            "indice":         g("indice"),
            "taxaCompra":     br_num(g("taxaCompra")),
            "taxaVenda":      br_num(g("taxaVenda")),
            "taxaIndicativa": br_num(g("taxaIndicativa")),
            "desvioPadrao":   br_num(g("desvioPadrao")),
            "pu":             br_num(g("pu")),
            "puPar":          br_num(g("puPar")),
            "duration":       br_num(g("duration")),
        })

    if ref_date is None:
        ref_date = _extract_date_from_filename(filepath)

    return ref_date, rows


def scan_directory(directory, parse_fn, extensions):
    """Scan a directory and parse all matching files. Returns {date: [rows]}."""
    history = {}
    if not os.path.isdir(directory):
        return history
    for fname in sorted(os.listdir(directory)):
        ext = os.path.splitext(fname)[1].lower()
        if ext not in extensions:
            continue
        fpath = os.path.join(directory, fname)
        try:
            _, rows = parse_fn(fpath)
            # Always use the filename date as the key — it is authoritative and
            # avoids collisions when the internal reference date is wrong.
            date_str = _extract_date_from_filename(fpath)
            if date_str and rows:
                history[date_str] = rows
        except Exception as e:
            print(f"  Warning: could not parse {fname}: {e}")
    return history
